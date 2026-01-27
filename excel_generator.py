#!/usr/bin/env python3
"""
Excel Maliyet Tablosu Generator

Generates company-format Excel cost tables from trafo_cli.py JSON results.
Matches the exact format of the reference TRF.26.0005 Excel file.

Usage:
    python excel_generator.py --input result.json --output maliyet.xlsx
    python excel_generator.py --input trafo1.json trafo2.json --output maliyet.xlsx
    python excel_generator.py --input results/ --output maliyet.xlsx --customer "ELSAN"
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Tuple

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Error: openpyxl is required. Install with: pip install openpyxl")
    sys.exit(1)


# Column widths matching reference file
TRAFO_COLUMN_WIDTHS = {
    'A': 68.57,
    'B': 12.86,
    'C': 14.71,
    'D': 30.71,
    'E': 13.0,
    'F': 29.43,
    'G': 29.29,
    'H': 11.86,
}

SUMMARY_COLUMN_WIDTHS = {
    'A': 29.29,
    'B': 12.0,
    'C': 27.14,
    'D': 30.71,
    'E': 23.0,
    'F': 13.0,
    'G': 16.57,
    'H': 15.71,
}

# Headers
TRAFO_HEADERS = [
    'MALZEME ADI',
    'MİKTAR',
    'BİRİM ',
    'YURTİÇİ \nHAMMADDE MALİYETİ',
    'YURTDIŞI \nHAMMADDE MALİYETİ',
    'YURTİÇİ \nMALİYET',
    'YURTDIŞI \nMALİYET',
    'SİPARİŞ ADEDİ'
]

SUMMARY_HEADERS = [
    'MALZEME TANIMI',
    'AĞIRLIK',
    'YURTİÇİ \nHAMMADDE MALİYETİ',
    'YURTDIŞI \nHAMMADDE MALİYETİ',
    'YURTİÇİ \nTRAFO MALİYETİ',
    'YURTDIŞI \nTRAFO MALİYETİ',
    'YURTİÇİ HAMMADDE YÜZDE DAĞILIM',
    'YURTDIŞI HAMMADDE YÜZDE DAĞILIM'
]


def create_styles() -> Dict:
    """Create style definitions."""
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    return {
        'header_font': Font(bold=True, size=11),
        'header_alignment': Alignment(horizontal='center', vertical='center', wrap_text=True),
        'header_fill': PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid'),
        'title_font': Font(bold=True, size=11),
        'title_fill': PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid'),
        'data_font': Font(size=10),
        'data_alignment': Alignment(horizontal='left', vertical='center'),
        'number_alignment': Alignment(horizontal='right', vertical='center'),
        'total_font': Font(bold=True, size=11),
        'total_fill': PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid'),
        'border': thin_border,
    }


def set_column_widths(ws, widths: Dict[str, float]):
    """Set column widths for worksheet."""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def apply_header_style(cell, styles: Dict):
    """Apply header style to a cell."""
    cell.font = styles['header_font']
    cell.alignment = styles['header_alignment']
    cell.fill = styles['header_fill']
    cell.border = styles['border']


def apply_title_style(cell, styles: Dict):
    """Apply title style to a cell."""
    cell.font = styles['title_font']
    cell.fill = styles['title_fill']
    cell.border = styles['border']


def apply_data_style(cell, styles: Dict, is_number: bool = False):
    """Apply data style to a cell."""
    cell.font = styles['data_font']
    cell.alignment = styles['number_alignment'] if is_number else styles['data_alignment']
    cell.border = styles['border']


def apply_total_style(cell, styles: Dict):
    """Apply total style to a cell."""
    cell.font = styles['total_font']
    cell.alignment = styles['number_alignment']
    cell.fill = styles['total_fill']
    cell.border = styles['border']


def get_conductor_name(trafo_data: Dict, conductor_type: Dict, is_lv: bool) -> str:
    """Get conductor name based on type and meta flags."""
    meta = trafo_data.get('meta', {})

    if is_lv:
        # LV conductor - check conductor_type.lv
        lv_type = conductor_type.get('lv', 'Folyo')
        return lv_type
    else:
        # HV conductor - check meta.circular_wire
        circular_wire = meta.get('circular_wire', False)
        if circular_wire:
            return "Yuvarlak Tel"
        else:
            # Check conductor_type.hv or default to Emaye Tel
            hv_type = conductor_type.get('hv', 'Emaye Tel')
            return hv_type


def get_trafo_title(order_info: Dict) -> str:
    """Generate transformer title from order info."""
    power = order_info.get('power_kva', 0)
    hv = order_info.get('hv_voltage', 0)
    lv = order_info.get('lv_voltage', 0)
    standard = order_info.get('standard', 'IEC')

    # Format voltage - if HV has secondary tap, show it
    hv_str = str(hv)
    if order_info.get('hv_secondary'):
        hv_str = f"{hv}({order_info['hv_secondary']})"

    return f"{power} kVA {hv_str} / {lv}  kV TRANSFORMATÖR  -  {standard}"


def write_trafo_sheet(wb: Workbook, trafo_list: List[Dict], order_infos: List[Dict],
                      prices_list: List[Dict], styles: Dict) -> Dict:
    """
    Write the TRAFO sheet with all transformers.
    Returns info needed for summary sheet.
    """
    ws = wb.active
    ws.title = 'TRAFO'

    set_column_widths(ws, TRAFO_COLUMN_WIDTHS)

    # Write headers in row 1
    for col, header in enumerate(TRAFO_HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        apply_header_style(cell, styles)

    ws.row_dimensions[1].height = 30

    current_row = 2
    trafo_info = {
        'blocks': [],  # List of (title_row, data_start, data_end, subtotal_row)
    }

    for i, (trafo_data, order_info, prices) in enumerate(zip(trafo_list, order_infos, prices_list)):
        block_info = write_trafo_block(
            ws, current_row, trafo_data, order_info, prices, styles
        )
        trafo_info['blocks'].append(block_info)
        current_row = block_info['next_row']

    # Grand total row
    current_row += 1

    # Build formula for total (multiply each subtotal by its quantity)
    if len(trafo_info['blocks']) == 1:
        block = trafo_info['blocks'][0]
        total_formula_f = f"=F{block['subtotal_row']}*$H${block['title_row']}"
        total_formula_g = f"=G{block['subtotal_row']}*$H${block['title_row']}"
    else:
        parts_f = []
        parts_g = []
        for block in trafo_info['blocks']:
            parts_f.append(f"F{block['subtotal_row']}*$H${block['title_row']}")
            parts_g.append(f"G{block['subtotal_row']}*$H${block['title_row']}")
        total_formula_f = "=" + "+".join(parts_f)
        total_formula_g = "=" + "+".join(parts_g)

    ws.cell(row=current_row, column=4, value='TOPLAM TEKLİF MALİYETİ')
    apply_total_style(ws.cell(row=current_row, column=4), styles)
    ws.cell(row=current_row, column=4).alignment = Alignment(horizontal='right', vertical='center')

    ws.cell(row=current_row, column=6, value=total_formula_f)
    apply_total_style(ws.cell(row=current_row, column=6), styles)

    ws.cell(row=current_row, column=7, value=total_formula_g)
    apply_total_style(ws.cell(row=current_row, column=7), styles)

    trafo_info['total_row'] = current_row
    return trafo_info


def write_trafo_block(ws, start_row: int, trafo_data: Dict, order_info: Dict,
                      prices: Dict, styles: Dict) -> Dict:
    """
    Write a single transformer cost block to the worksheet.
    Returns dict with row info.
    """
    material = order_info.get('material', 'Bakır')
    quantity = order_info.get('quantity', 1)
    labor_hours = order_info.get('labor_hours', 0)
    conductor_type = order_info.get('conductor_type', {'lv': 'Folyo', 'hv': 'Emaye Tel'})

    # Extract data from trafo_data
    cost = trafo_data.get('cost', {})
    tank_oil = trafo_data.get('tank_oil', {})

    foil_weight = cost.get('foil_weight', 0)
    wire_weight = cost.get('wire_weight', 0)
    core_weight = cost.get('core_weight', 0)
    tank_weight = tank_oil.get('tank', {}).get('total_weight', 0)
    oil_weight = tank_oil.get('oil', {}).get('weight', 0)

    # Prices
    lv_price = prices.get('lv_conductor_per_kg', 0)
    hv_price = prices.get('hv_conductor_per_kg', 0)
    core_price = prices.get('core_per_kg', 0)
    mech_price = prices.get('mechanical_per_kg', 0)
    oil_price = prices.get('oil_per_kg', 0)
    labor_rate = prices.get('labor_per_hour', 0)
    accessory_fixed = prices.get('accessory_fixed', 0)
    other_fixed = prices.get('other_fixed', 0)

    # Conductor names
    lv_conductor = get_conductor_name(trafo_data, conductor_type, is_lv=True)
    yg_conductor = get_conductor_name(trafo_data, conductor_type, is_lv=False)

    row = start_row
    title_row = row

    # Title row
    title = get_trafo_title(order_info)
    cell = ws.cell(row=row, column=1, value=title)
    apply_title_style(cell, styles)
    for col in range(2, 8):
        apply_title_style(ws.cell(row=row, column=col), styles)
    cell = ws.cell(row=row, column=8, value=quantity)
    apply_title_style(cell, styles)
    row += 1

    data_start_row = row

    # AG İLETKEN
    ws.cell(row=row, column=1, value=f"AG İLETKEN - {lv_conductor} - {material}")
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value=foil_weight)
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)
    ws.cell(row=row, column=3, value='kg')
    apply_data_style(ws.cell(row=row, column=3), styles)
    ws.cell(row=row, column=4, value=lv_price)
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
    ws.cell(row=row, column=5, value=lv_price)
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    ws.cell(row=row, column=7, value=f"=B{row}*E{row}")
    apply_data_style(ws.cell(row=row, column=7), styles, is_number=True)
    row += 1

    # YG İLETKEN
    ws.cell(row=row, column=1, value=f"YG İLETKEN - {yg_conductor} - {material}")
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value=wire_weight)
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)
    ws.cell(row=row, column=3, value='kg')
    apply_data_style(ws.cell(row=row, column=3), styles)
    ws.cell(row=row, column=4, value=hv_price)
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
    ws.cell(row=row, column=5, value=hv_price)
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    ws.cell(row=row, column=7, value=f"=B{row}*E{row}")
    apply_data_style(ws.cell(row=row, column=7), styles, is_number=True)
    row += 1

    # NÜVE
    ws.cell(row=row, column=1, value='NÜVE 27-110')
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value=core_weight)
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)
    ws.cell(row=row, column=3, value='kg')
    apply_data_style(ws.cell(row=row, column=3), styles)
    ws.cell(row=row, column=4, value=core_price)
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
    ws.cell(row=row, column=5, value=core_price)
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    ws.cell(row=row, column=7, value=f"=B{row}*E{row}")
    apply_data_style(ws.cell(row=row, column=7), styles, is_number=True)
    row += 1

    # MEKANİK AKSAM
    ws.cell(row=row, column=1, value='MEKANİK AKSAM')
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value=tank_weight)
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)
    ws.cell(row=row, column=3, value='kg')
    apply_data_style(ws.cell(row=row, column=3), styles)
    ws.cell(row=row, column=4, value=mech_price)
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
    ws.cell(row=row, column=5, value=mech_price)
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    ws.cell(row=row, column=7, value=f"=B{row}*E{row}")
    apply_data_style(ws.cell(row=row, column=7), styles, is_number=True)
    row += 1

    # YAĞ
    ws.cell(row=row, column=1, value='YAĞ')
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value=oil_weight)
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)
    ws.cell(row=row, column=3, value='kg')
    apply_data_style(ws.cell(row=row, column=3), styles)
    ws.cell(row=row, column=4, value=oil_price)
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
    ws.cell(row=row, column=5, value=oil_price)
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    ws.cell(row=row, column=7, value=f"=B{row}*E{row}")
    apply_data_style(ws.cell(row=row, column=7), styles, is_number=True)
    row += 1

    # İŞÇİLİK
    ws.cell(row=row, column=1, value='İŞÇİLİK')
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value=labor_hours)
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)
    ws.cell(row=row, column=3, value='adam/saat')
    apply_data_style(ws.cell(row=row, column=3), styles)
    ws.cell(row=row, column=4, value=labor_rate)
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
    ws.cell(row=row, column=5, value=labor_rate)
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    ws.cell(row=row, column=7, value=f"=B{row}*E{row}")
    apply_data_style(ws.cell(row=row, column=7), styles, is_number=True)
    row += 1

    # AKSESUAR
    ws.cell(row=row, column=1, value='AKSESUAR')
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value=1)
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)
    ws.cell(row=row, column=3, value='-')
    apply_data_style(ws.cell(row=row, column=3), styles)
    ws.cell(row=row, column=4, value=accessory_fixed)
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
    ws.cell(row=row, column=5, value=accessory_fixed)
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    ws.cell(row=row, column=7, value=f"=B{row}*E{row}")
    apply_data_style(ws.cell(row=row, column=7), styles, is_number=True)
    row += 1

    # DİĞER
    ws.cell(row=row, column=1, value='DİĞER (PRESPAN, TAKOZ VS.)')
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value=1)
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)
    ws.cell(row=row, column=3, value='-')
    apply_data_style(ws.cell(row=row, column=3), styles)
    ws.cell(row=row, column=4, value=other_fixed)
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
    ws.cell(row=row, column=5, value=other_fixed)
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    ws.cell(row=row, column=7, value=f"=B{row}*E{row}")
    apply_data_style(ws.cell(row=row, column=7), styles, is_number=True)

    data_end_row = row
    row += 1

    # GARANTİ PAYI
    garanti_row = row
    ws.cell(row=row, column=1, value='GARANTİ PAYI')
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value='-')
    apply_data_style(ws.cell(row=row, column=2), styles)
    ws.cell(row=row, column=3, value='-')
    apply_data_style(ws.cell(row=row, column=3), styles)
    ws.cell(row=row, column=4, value='-')
    apply_data_style(ws.cell(row=row, column=4), styles)
    ws.cell(row=row, column=5, value='-')
    apply_data_style(ws.cell(row=row, column=5), styles)
    ws.cell(row=row, column=6, value=0.05)
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    ws.cell(row=row, column=7, value=0.05)
    apply_data_style(ws.cell(row=row, column=7), styles, is_number=True)
    row += 1

    # Subtotal row: =SUM(F3:F10)*F11+SUM(F3:F10)
    subtotal_row = row
    for col in range(1, 6):
        apply_data_style(ws.cell(row=row, column=col), styles)

    ws.cell(row=row, column=6, value=f"=SUM(F{data_start_row}:F{data_end_row})*F{garanti_row}+SUM(F{data_start_row}:F{data_end_row})")
    apply_total_style(ws.cell(row=row, column=6), styles)

    ws.cell(row=row, column=7, value=f"=SUM(G{data_start_row}:G{data_end_row})*G{garanti_row}+SUM(G{data_start_row}:G{data_end_row})")
    apply_total_style(ws.cell(row=row, column=7), styles)

    return {
        'title_row': title_row,
        'data_start_row': data_start_row,
        'data_end_row': data_end_row,
        'garanti_row': garanti_row,
        'subtotal_row': subtotal_row,
        'next_row': row + 2,  # +2 for blank row spacing
        'conductor_type': {
            'lv': lv_conductor,
            'hv': yg_conductor
        }
    }


def write_summary_sheet(wb: Workbook, trafo_info: Dict, trafo_list: List[Dict],
                        order_infos: List[Dict], prices_list: List[Dict], styles: Dict):
    """Write the MALİYET ÖZETİ summary sheet with formulas referencing TRAFO sheet."""
    ws = wb.create_sheet('MALİYET ÖZETİ')

    set_column_widths(ws, SUMMARY_COLUMN_WIDTHS)

    # Write headers in row 2 (row 1 is empty in reference)
    for col, header in enumerate(SUMMARY_HEADERS, 1):
        cell = ws.cell(row=2, column=col, value=header)
        apply_header_style(cell, styles)

    ws.row_dimensions[2].height = 30

    blocks = trafo_info['blocks']

    # Collect conductor types used across all transformers
    lv_types = set()
    hv_types = set()
    for i, block in enumerate(blocks):
        lv_types.add(block['conductor_type']['lv'])
        hv_types.add(block['conductor_type']['hv'])

    row = 3

    # Build weight formulas that reference TRAFO sheet
    # Format: =TRAFO!B3*TRAFO!$H$2 + TRAFO!B14*TRAFO!$H$13 ...

    def build_weight_formula(row_offset: int) -> str:
        """Build formula to sum weights from all trafo blocks with quantity multiplier."""
        parts = []
        for block in blocks:
            data_row = block['data_start_row'] + row_offset
            title_row = block['title_row']
            parts.append(f"TRAFO!B{data_row}*TRAFO!$H${title_row}")
        return "=" + "+".join(parts)

    def build_single_weight_formula(block_idx: int, row_offset: int) -> str:
        """Build formula for a single block's weight with quantity."""
        block = blocks[block_idx]
        data_row = block['data_start_row'] + row_offset
        title_row = block['title_row']
        return f"=TRAFO!B{data_row}*TRAFO!$H${title_row}"

    # Determine which rows have which conductor types
    # We need to group by conductor type similar to reference file

    # AG İletken rows (one per LV type found)
    for lv_type in sorted(lv_types):
        ws.cell(row=row, column=1, value=f"AG İletken - {lv_type}")
        apply_data_style(ws.cell(row=row, column=1), styles)

        # Build formula for blocks using this LV type
        parts = []
        for i, block in enumerate(blocks):
            if block['conductor_type']['lv'] == lv_type:
                data_row = block['data_start_row']  # AG is first row
                title_row = block['title_row']
                parts.append(f"TRAFO!B{data_row}*TRAFO!$H${title_row}")

        if parts:
            ws.cell(row=row, column=2, value="=" + "+".join(parts))
        else:
            ws.cell(row=row, column=2, value=0)
        apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)

        # Get price from first matching transformer
        lv_price = 0
        for i, block in enumerate(blocks):
            if block['conductor_type']['lv'] == lv_type:
                lv_price = prices_list[i].get('lv_conductor_per_kg', 0)
                break

        ws.cell(row=row, column=3, value=lv_price)
        apply_data_style(ws.cell(row=row, column=3), styles, is_number=True)
        ws.cell(row=row, column=4, value=lv_price)
        apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
        ws.cell(row=row, column=5, value=f"=B{row}*C{row}")
        apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
        ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
        apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
        row += 1

    # YG İletken rows (one per HV type found)
    for hv_type in sorted(hv_types):
        ws.cell(row=row, column=1, value=f"YG İletken - İletken {hv_type}")
        apply_data_style(ws.cell(row=row, column=1), styles)

        # Build formula for blocks using this HV type
        parts = []
        for i, block in enumerate(blocks):
            if block['conductor_type']['hv'] == hv_type:
                data_row = block['data_start_row'] + 1  # YG is second row
                title_row = block['title_row']
                parts.append(f"TRAFO!B{data_row}*TRAFO!$H${title_row}")

        if parts:
            ws.cell(row=row, column=2, value="=" + "+".join(parts))
        else:
            ws.cell(row=row, column=2, value=0)
        apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)

        # Get price from first matching transformer
        hv_price = 0
        for i, block in enumerate(blocks):
            if block['conductor_type']['hv'] == hv_type:
                hv_price = prices_list[i].get('hv_conductor_per_kg', 0)
                break

        ws.cell(row=row, column=3, value=hv_price)
        apply_data_style(ws.cell(row=row, column=3), styles, is_number=True)
        ws.cell(row=row, column=4, value=hv_price)
        apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
        ws.cell(row=row, column=5, value=f"=B{row}*C{row}")
        apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
        ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
        apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
        row += 1

    # Nüve
    ws.cell(row=row, column=1, value='Nüve')
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value=build_weight_formula(2))  # NÜVE is 3rd row (offset 2)
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)
    core_price = prices_list[0].get('core_per_kg', 0)
    ws.cell(row=row, column=3, value=core_price)
    apply_data_style(ws.cell(row=row, column=3), styles, is_number=True)
    ws.cell(row=row, column=4, value=core_price)
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
    ws.cell(row=row, column=5, value=f"=B{row}*C{row}")
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    row += 1

    # Mekanik Aksam
    ws.cell(row=row, column=1, value='Mekanik Aksam')
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value=build_weight_formula(3))  # MEKANİK is 4th row (offset 3)
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)
    mech_price = prices_list[0].get('mechanical_per_kg', 0)
    ws.cell(row=row, column=3, value=mech_price)
    apply_data_style(ws.cell(row=row, column=3), styles, is_number=True)
    ws.cell(row=row, column=4, value=mech_price)
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
    ws.cell(row=row, column=5, value=f"=B{row}*C{row}")
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    row += 1

    # Yağ
    ws.cell(row=row, column=1, value='Yağ')
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value=build_weight_formula(4))  # YAĞ is 5th row (offset 4)
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)
    oil_price = prices_list[0].get('oil_per_kg', 0)
    ws.cell(row=row, column=3, value=oil_price)
    apply_data_style(ws.cell(row=row, column=3), styles, is_number=True)
    ws.cell(row=row, column=4, value=oil_price)
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
    ws.cell(row=row, column=5, value=f"=B{row}*C{row}")
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    row += 1

    # İşçilik
    ws.cell(row=row, column=1, value='İşçilik')
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value=build_weight_formula(5))  # İŞÇİLİK is 6th row (offset 5)
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)
    labor_rate = prices_list[0].get('labor_per_hour', 0)
    ws.cell(row=row, column=3, value=labor_rate)
    apply_data_style(ws.cell(row=row, column=3), styles, is_number=True)
    ws.cell(row=row, column=4, value=labor_rate)
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)
    ws.cell(row=row, column=5, value=f"=B{row}*C{row}")
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=B{row}*D{row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    row += 1

    # Aksesuarlar - uses formula referencing TRAFO sheet prices
    ws.cell(row=row, column=1, value='Aksesuarlar')
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value="=1")
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)

    # Build accessory price formula
    acc_parts = []
    for block in blocks:
        acc_row = block['data_start_row'] + 6  # AKSESUAR is 7th row (offset 6)
        title_row = block['title_row']
        acc_parts.append(f"TRAFO!$H${title_row}*TRAFO!D{acc_row}")
    ws.cell(row=row, column=3, value="=" + "+".join(acc_parts))
    apply_data_style(ws.cell(row=row, column=3), styles, is_number=True)

    acc_parts_e = []
    for block in blocks:
        acc_row = block['data_start_row'] + 6
        title_row = block['title_row']
        acc_parts_e.append(f"TRAFO!$H${title_row}*TRAFO!E{acc_row}")
    ws.cell(row=row, column=4, value="=" + "+".join(acc_parts_e))
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)

    ws.cell(row=row, column=5, value=f"=C{row}*$B${row}")
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=D{row}*$B${row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)
    row += 1

    # Diğer
    ws.cell(row=row, column=1, value='Diğer (prespan, takoz vs.)')
    apply_data_style(ws.cell(row=row, column=1), styles)
    ws.cell(row=row, column=2, value="=1")
    apply_data_style(ws.cell(row=row, column=2), styles, is_number=True)

    # Build other price formula
    other_parts = []
    for block in blocks:
        other_row = block['data_start_row'] + 7  # DİĞER is 8th row (offset 7)
        title_row = block['title_row']
        other_parts.append(f"TRAFO!$H${title_row}*TRAFO!D{other_row}")
    ws.cell(row=row, column=3, value="=" + "+".join(other_parts))
    apply_data_style(ws.cell(row=row, column=3), styles, is_number=True)

    other_parts_e = []
    for block in blocks:
        other_row = block['data_start_row'] + 7
        title_row = block['title_row']
        other_parts_e.append(f"TRAFO!$H${title_row}*TRAFO!E{other_row}")
    ws.cell(row=row, column=4, value="=" + "+".join(other_parts_e))
    apply_data_style(ws.cell(row=row, column=4), styles, is_number=True)

    ws.cell(row=row, column=5, value=f"=C{row}*$B${row}")
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=f"=D{row}*$B${row}")
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)

    data_end_row = row
    row += 1

    # Garanti Payı
    ws.cell(row=row, column=1, value='Garanti Payı')
    apply_data_style(ws.cell(row=row, column=1), styles)
    apply_data_style(ws.cell(row=row, column=2), styles)
    ws.cell(row=row, column=3, value='-')
    apply_data_style(ws.cell(row=row, column=3), styles)
    ws.cell(row=row, column=4, value='-')
    apply_data_style(ws.cell(row=row, column=4), styles)
    ws.cell(row=row, column=5, value=0.05)
    apply_data_style(ws.cell(row=row, column=5), styles, is_number=True)
    ws.cell(row=row, column=6, value=0.05)
    apply_data_style(ws.cell(row=row, column=6), styles, is_number=True)

    garanti_row = row
    row += 1

    # Total row
    total_row = row
    ws.cell(row=row, column=5, value=f"=SUM(E3:E{data_end_row})*E{garanti_row}+SUM(E3:E{data_end_row})")
    apply_total_style(ws.cell(row=row, column=5), styles)
    ws.cell(row=row, column=6, value=f"=SUM(F3:F{data_end_row})*F{garanti_row}+SUM(F3:F{data_end_row})")
    apply_total_style(ws.cell(row=row, column=6), styles)

    # Now add percentage formulas (columns G and H)
    for r in range(3, data_end_row + 1):
        ws.cell(row=r, column=7, value=f"=E{r}/$E${total_row}")
        apply_data_style(ws.cell(row=r, column=7), styles, is_number=True)
        ws.cell(row=r, column=7).number_format = '0.00%'

        ws.cell(row=r, column=8, value=f"=F{r}/$F${total_row}")
        apply_data_style(ws.cell(row=r, column=8), styles, is_number=True)
        ws.cell(row=r, column=8).number_format = '0.00%'


def load_json_files(input_paths: List[str]) -> List[Dict]:
    """Load JSON data from files or directory."""
    json_data = []

    for path in input_paths:
        p = Path(path)
        if p.is_dir():
            for json_file in sorted(p.glob('*.json')):
                with open(json_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    json_data.append(data)
        elif p.is_file():
            with open(p, 'r', encoding='utf-8') as f:
                data = json.load(f)
                json_data.append(data)
        else:
            print(f"Warning: Path not found: {path}", file=sys.stderr)

    return json_data


def parse_combined_json(data: Dict) -> Tuple[Dict, Dict, Dict]:
    """
    Parse a combined JSON that includes trafo data, order_info, and prices.
    Returns (trafo_data, order_info, prices) tuple.
    """
    if 'order_info' in data:
        order_info = data['order_info']
        prices = data.get('prices', {})
        trafo_data = {k: v for k, v in data.items() if k not in ['order_info', 'prices']}
        return trafo_data, order_info, prices
    else:
        trafo_data = data
        order_info = {
            'power_kva': 0,
            'hv_voltage': 0,
            'lv_voltage': 0,
            'quantity': 1,
            'standard': 'IEC',
            'customer': 'UNKNOWN',
            'labor_hours': 0,
            'material': 'Bakır',
            'conductor_type': {'lv': 'Folyo', 'hv': 'Emaye Tel'}
        }
        prices = {
            'lv_conductor_per_kg': 0,
            'hv_conductor_per_kg': 0,
            'core_per_kg': 0,
            'mechanical_per_kg': 0,
            'oil_per_kg': 0,
            'labor_per_hour': 0,
            'accessory_fixed': 0,
            'other_fixed': 0
        }
        return trafo_data, order_info, prices


def generate_filename(customer: Optional[str], sequence: int = 1) -> str:
    """Generate output filename in format TRF_{year}_{sequence}_{customer}.xlsx"""
    year = datetime.now().year
    year_short = str(year)[2:]  # Last 2 digits
    customer_part = customer.replace(' ', '_') if customer else 'TRAFO'
    return f"TRF.{year_short}.{sequence:04d} - {customer_part}.xlsx"


def generate_excel(input_files: List[str], output_path: Optional[str], customer: Optional[str] = None):
    """Main function to generate Excel cost table."""
    json_data_list = load_json_files(input_files)

    if not json_data_list:
        print("Error: No JSON data loaded", file=sys.stderr)
        sys.exit(1)

    trafo_list = []
    order_infos = []
    prices_list = []

    for data in json_data_list:
        trafo_data, order_info, prices = parse_combined_json(data)
        trafo_list.append(trafo_data)
        order_infos.append(order_info)
        prices_list.append(prices)

    if customer:
        for order_info in order_infos:
            order_info['customer'] = customer

    wb = Workbook()
    styles = create_styles()

    trafo_info = write_trafo_sheet(wb, trafo_list, order_infos, prices_list, styles)
    write_summary_sheet(wb, trafo_info, trafo_list, order_infos, prices_list, styles)

    if not output_path:
        first_customer = order_infos[0].get('customer', customer) if order_infos else customer
        output_path = generate_filename(first_customer or 'TRAFO')

    wb.save(output_path)
    print(f"Excel file generated: {output_path}")

    return output_path


def main():
    parser = argparse.ArgumentParser(
        description='Generate Excel cost tables from trafo_cli.py JSON results.',
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog='''
Examples:
  python excel_generator.py --input result.json --output maliyet.xlsx
  python excel_generator.py --input trafo1.json trafo2.json --output maliyet.xlsx
  python excel_generator.py --input results/ --output maliyet.xlsx --customer "ELSAN"
        '''
    )

    parser.add_argument(
        '--input', '-i',
        nargs='+',
        required=True,
        help='Input JSON file(s) or directory containing JSON files'
    )

    parser.add_argument(
        '--output', '-o',
        help='Output Excel file path (default: auto-generated)'
    )

    parser.add_argument(
        '--customer', '-c',
        help='Customer name (overrides value in JSON)'
    )

    args = parser.parse_args()

    generate_excel(args.input, args.output, args.customer)


if __name__ == '__main__':
    main()
